"""
ALPER-EU — Session Data Exporter
==================================
Generates a teacher-friendly Excel workbook from a session analysis folder.

Sheets produced:
  1. Session Report      — plain-English summary of the session
  2. Engagement Summary  — one row per second, score + engagement level
  3. Gaze Data           — gaze coordinates + on-screen status per frame
  4. Body Data           — posture / distance / facing per frame
  5. Disengagement Events — each disengagement event with timing

Usage:
    from session_exporter import export_session_xlsx
    path = export_session_xlsx(session_folder, output_path)

Or from terminal:
    python analysis/session_exporter.py recordings/both_20260306_110143
"""

import json
import os
import sys
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_NAME = 'Arial'

# Header colours
CLR_HEADER_DARK   = '1F3864'   # dark navy  — sheet title rows
CLR_HEADER_MID    = '2E75B6'   # mid blue   — column headers
CLR_HEADER_LIGHT  = 'D6E4F0'   # pale blue  — sub-headers / alternating rows
CLR_GREEN_DARK    = '1E8449'
CLR_GREEN_LIGHT   = 'D5F5E3'
CLR_AMBER_DARK    = 'D35400'
CLR_AMBER_LIGHT   = 'FDEBD0'
CLR_RED_DARK      = 'C0392B'
CLR_RED_LIGHT     = 'FADBD8'
CLR_WHITE         = 'FFFFFF'
CLR_GREY_ROW      = 'F2F2F2'

THIN = Side(style='thin', color='BFBFBF')
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _font(bold=False, size=10, color='000000', name=FONT_NAME):
    return Font(name=name, bold=bold, size=size, color=color)


def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_row(ws, row, values, bg=CLR_HEADER_MID, fg=CLR_WHITE,
                bold=True, size=10):
    """Write a styled header row."""
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=bold, size=size, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align(h='center')
        c.border    = BORDER_THIN


def _data_row(ws, row, values, alternate=False, formats=None):
    """Write a data row with optional alternating row shading."""
    bg = CLR_GREY_ROW if alternate else CLR_WHITE
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font()
        c.fill      = _fill(bg)
        c.alignment = _align(h='center')
        c.border    = BORDER_THIN
        if formats and col <= len(formats) and formats[col - 1]:
            c.number_format = formats[col - 1]


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_title(ws, title, subtitle, col_span):
    """Write a two-row title block at the top of a sheet."""
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=col_span)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = _font(bold=True, size=13, color=CLR_WHITE)
    c.fill      = _fill(CLR_HEADER_DARK)
    c.alignment = _align(h='center')

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=col_span)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font      = _font(size=9, color='595959')
    c.fill      = _fill('EBF3FB')
    c.alignment = _align(h='center')
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16


def _score_fill(score):
    """Return fill colour based on engagement score zone."""
    if score >= 80:
        return _fill(CLR_GREEN_LIGHT)
    elif score >= 50:
        return _fill(CLR_AMBER_LIGHT)
    else:
        return _fill(CLR_RED_LIGHT)


def _score_label(score):
    if score >= 80:
        return 'High'
    elif score >= 50:
        return 'Moderate'
    else:
        return 'Low'


# ---------------------------------------------------------------------------
# Sheet 1 — Session Report
# ---------------------------------------------------------------------------
def _write_session_report(wb, summary, session_label, metadata=None):
    ws = wb.active
    ws.title = 'Session Report'

    _sheet_title(ws, 'ALPER-EU — Session Engagement Report',
                 f'Session: {session_label}   |   '
                 f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}',
                 col_span=3)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 48

    # Load metadata if not passed in (read from session folder)
    if metadata is None:
        meta_path = os.path.join(
            os.path.dirname(summary.get('session_folder', '')),
            os.path.basename(summary.get('session_folder', '')),
            'session_metadata.json')
        # Also try analysis-level path resolution
        if not os.path.exists(meta_path):
            sf = summary.get('session_folder', '')
            meta_path = os.path.join(sf, 'session_metadata.json')
        if os.path.exists(meta_path):
            try:
                with open(meta_path) as f:
                    metadata = json.load(f)
            except Exception:
                metadata = None

    meta_section = []
    if metadata:
        meta_section = [
            ('Student ID',
             metadata.get('student_id', '—'),
             'Identifier for the student in this session'),
            ('Task name',
             metadata.get('task_name', '—'),
             'The learning task the student was performing'),
            ('Facilitator',
             metadata.get('facilitator', '—') or '—',
             'Teacher or researcher who ran the session'),
            ('Session notes',
             metadata.get('notes', '—') or '—',
             'Free-text notes recorded at the start of the session'),
            ('Session type',
             'Guest / Demo' if metadata.get('guest') else 'Recorded',
             'Whether the session used the guest (no metadata) mode'),
            ('Recorded at',
             metadata.get('recorded_at', '—'),
             'ISO timestamp when the session started'),
        ]

    sections = [
        # (Section heading, [(label, value, explanation), ...])
        *([('Participant & Task', meta_section)] if meta_section else []),
        ('Session Overview', [
            ('Session folder',
             summary.get('session_folder', '—'),
             'The folder where all raw data is stored'),
            ('Duration',
             f"{summary.get('duration_s', 0):.1f} seconds",
             'Total length of the recording session'),
            ('Total frames analysed',
             f"{summary.get('total_frames', 0):,}",
             'Number of body tracking frames processed'),
            ('Frame rate',
             f"{summary.get('effective_fps', 0):.1f} FPS",
             'Frames captured per second (target: 30)'),
        ]),
        ('Engagement Results', [
            ('Mean engagement score',
             f"{summary.get('mean_engagement_score', '—')} / 100",
             'Average score across the session. '
             'High ≥80  |  Moderate 50–79  |  Low <50'),
            ('Minimum score',
             f"{summary.get('min_engagement_score', '—')} / 100",
             'Lowest engagement score recorded (any 1-second window)'),
            ('Maximum score',
             f"{summary.get('max_engagement_score', '—')} / 100",
             'Highest engagement score recorded'),
            ('Time engaged',
             f"{summary.get('engaged_pct', 0):.1f}%",
             'Percentage of session where student was engaged '
             '(at least one signal active)'),
            ('Time disengaged',
             f"{summary.get('disengaged_pct', 0):.1f}%",
             'Percentage of session where ALL THREE signals failed '
             'simultaneously'),
        ]),
        ('Signal Breakdown', [
            ('Gaze on-screen',
             f"{summary.get('gaze_on_screen_pct', 0):.1f}%",
             'How often the student was looking at the screen '
             '(weight: 50% of score)'),
            ('Facing forward',
             f"{summary.get('facing_forward_pct', 0):.1f}%",
             'How often the student was facing the screen '
             '(weight: 30% of score)'),
            ('Distance OK',
             f"{summary.get('distance_ok_pct', 0):.1f}%",
             'How often the student was within appropriate distance '
             '(weight: 20% of score)'),
            ('Body engaged',
             f"{summary.get('body_engaged_pct', 0):.1f}%",
             'Combined posture signal (facing + distance both OK)'),
        ]),
        ('Disengagement Events', [
            ('Number of events',
             str(summary.get('disengagement_events', 0)),
             'How many times the student became fully disengaged'),
            ('Average duration',
             f"{summary.get('avg_disengagement_dur_s', 0):.2f} seconds",
             'Average length of each disengagement episode. '
             'Under 2s = natural task switching; over 5s = concern'),
        ]),
        ('Score Weights Used', [
            ('Gaze weight',     '50%', 'Contribution of gaze signal to score'),
            ('Facing weight',   '30%', 'Contribution of facing signal to score'),
            ('Distance weight', '20%', 'Contribution of distance signal to score'),
        ]),
    ]

    current_row = 4
    for section_title, items in sections:
        # Section header
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3)
        c = ws.cell(row=current_row, column=1, value=section_title)
        c.font      = _font(bold=True, size=11, color=CLR_WHITE)
        c.fill      = _fill(CLR_HEADER_MID)
        c.alignment = _align(h='left')
        c.border    = BORDER_THIN
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Column sub-headers
        for col, label in enumerate(['Metric', 'Value', 'What this means'],
                                    start=1):
            c = ws.cell(row=current_row, column=col, value=label)
            c.font      = _font(bold=True, size=9, color='595959')
            c.fill      = _fill(CLR_HEADER_LIGHT)
            c.alignment = _align(h='left')
            c.border    = BORDER_THIN
        current_row += 1

        for i, (label, value, explanation) in enumerate(items):
            alt = (i % 2 == 1)
            bg  = CLR_GREY_ROW if alt else CLR_WHITE

            for col, text in enumerate([label, value, explanation], start=1):
                c = ws.cell(row=current_row, column=col, value=text)
                c.font      = _font(bold=(col == 1), size=10)
                c.fill      = _fill(bg)
                c.alignment = _align(h='left', wrap=(col == 3))
                c.border    = BORDER_THIN
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1  # blank row between sections


# ---------------------------------------------------------------------------
# Sheet 2 — Engagement Summary (per second)
# ---------------------------------------------------------------------------
def _write_engagement_summary(wb, score_df):
    ws = wb.create_sheet('Engagement Summary')
    cols = 7
    _sheet_title(ws, 'Engagement Summary — Per Second',
                 'One row per second of the session. '
                 'Score: High ≥80  |  Moderate 50–79  |  Low <50',
                 col_span=cols)

    headers = ['Time (seconds)', 'Raw Score (0-100)', 'Smoothed Score (0-100)',
               'Engagement Level', 'Gaze On-Screen (%)',
               'Facing Forward (%)', 'Distance OK (%)']
    _header_row(ws, 3, headers)

    widths = [18, 20, 22, 18, 20, 20, 18]
    _set_col_widths(ws, widths)

    for i, (_, row) in enumerate(score_df.iterrows()):
        r     = i + 4
        score = row['score_smooth']
        level = _score_label(score)

        values = [
            row['t_s'],
            round(row['score'], 1),
            round(score, 1),
            level,
            round(row['gaze']     * 100, 1),
            round(row['facing']   * 100, 1),
            round(row['distance'] * 100, 1),
        ]

        alt = (i % 2 == 1)
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=10)
            c.alignment = _align(h='center')
            c.border    = BORDER_THIN

            # Colour the score and level cells by zone
            if col in (2, 3, 4):
                c.fill = _score_fill(score)
                if col == 4:
                    colour = (CLR_GREEN_DARK if score >= 80
                              else CLR_AMBER_DARK if score >= 50
                              else CLR_RED_DARK)
                    c.font = _font(bold=True, size=10, color=colour)
            else:
                c.fill = _fill(CLR_GREY_ROW if alt else CLR_WHITE)

        # Percentage formats for last 3 cols
        for col in (5, 6, 7):
            ws.cell(row=r, column=col).number_format = '0.0"%"'

    # Summary row at bottom
    n = len(score_df) + 4
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=2)
    ws.cell(row=n, column=1, value='Session Average').font = _font(bold=True)
    ws.cell(row=n, column=1).fill = _fill(CLR_HEADER_LIGHT)

    avg_col_map = {3: 'score_smooth', 5: 'gaze', 6: 'facing', 7: 'distance'}
    for col, field in avg_col_map.items():
        mult = 100 if field != 'score_smooth' else 1
        val = round(score_df[field].mean() * mult, 1)
        c = ws.cell(row=n, column=col, value=val)
        c.font = _font(bold=True)
        c.fill = _fill(CLR_HEADER_LIGHT)
        c.border = BORDER_THIN
        c.alignment = _align(h='center')


# ---------------------------------------------------------------------------
# Sheet 3 — Gaze Data
# ---------------------------------------------------------------------------
def _write_gaze_data(wb, merged_df):
    ws = wb.create_sheet('Gaze Data')
    cols = 6
    _sheet_title(ws, 'Gaze Tracking Data — Per Frame',
                 'Raw gaze coordinates captured by the Tobii Eye Tracker 4C. '
                 'Screen resolution: 1920 × 1200 pixels.',
                 col_span=cols)

    headers = ['Time (seconds)', 'Gaze X (pixels)', 'Gaze Y (pixels)',
               'On Screen', 'Screen Region', 'Engagement Verdict']
    _header_row(ws, 3, headers)
    _set_col_widths(ws, [16, 16, 16, 12, 18, 20])

    def _region(x, y, on_screen):
        """Divide screen into a 3×3 grid and label the region."""
        if not on_screen or pd.isna(x) or pd.isna(y):
            return 'Off Screen'
        col = 'Left' if x < 640 else ('Centre' if x < 1280 else 'Right')
        row = 'Top'  if y < 400 else ('Middle' if y < 800  else 'Bottom')
        return f'{row} {col}'

    for i, (_, row) in enumerate(merged_df.iterrows()):
        r          = i + 4
        on_screen  = bool(row.get('on_screen', False))
        x          = row.get('x', None)
        y          = row.get('y', None)
        engaged    = bool(row.get('Engaged', True))
        alt        = (i % 2 == 1)
        bg         = CLR_GREY_ROW if alt else CLR_WHITE

        values = [
            round(float(row['t']), 3),
            round(float(x), 1) if pd.notna(x) and on_screen else '—',
            round(float(y), 1) if pd.notna(y) and on_screen else '—',
            'Yes' if on_screen else 'No',
            _region(x, y, on_screen),
            'Engaged' if engaged else 'Disengaged',
        ]

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9)
            c.alignment = _align(h='center')
            c.border    = BORDER_THIN

            if col == 4:   # On Screen — colour green/red
                c.fill = _fill(CLR_GREEN_LIGHT if on_screen else CLR_RED_LIGHT)
                c.font = _font(bold=True, size=9,
                               color=CLR_GREEN_DARK if on_screen
                               else CLR_RED_DARK)
            elif col == 6:  # Engagement verdict
                c.fill = _fill(CLR_GREEN_LIGHT if engaged else CLR_RED_LIGHT)
                c.font = _font(bold=True, size=9,
                               color=CLR_GREEN_DARK if engaged
                               else CLR_RED_DARK)
            else:
                c.fill = _fill(bg)


# ---------------------------------------------------------------------------
# Sheet 4 — Body Data
# ---------------------------------------------------------------------------
def _write_body_data(wb, merged_df):
    ws = wb.create_sheet('Body Data')
    cols = 7
    _sheet_title(ws, 'Body Tracking Data — Per Frame',
                 'Posture and position data captured by the Intel RealSense D455.',
                 col_span=cols)

    headers = ['Time (seconds)', 'Distance (metres)', 'Distance OK',
               'Facing Forward', 'Body Engaged', 'Nose X (norm.)',
               'Nose Y (norm.)']
    _header_row(ws, 3, headers)
    _set_col_widths(ws, [16, 18, 14, 14, 14, 16, 16])

    for i, (_, row) in enumerate(merged_df.iterrows()):
        r        = i + 4
        dist     = row.get('Nose_Dist_M', None)
        dist_ok  = bool(row.get('Distance_OK', False))
        facing   = bool(row.get('Facing_Forward', False))
        engaged  = bool(row.get('Body_Engaged', False))
        alt      = (i % 2 == 1)
        bg       = CLR_GREY_ROW if alt else CLR_WHITE

        values = [
            round(float(row['t']), 3),
            round(float(dist), 3) if pd.notna(dist) else '—',
            'Yes' if dist_ok  else 'No',
            'Yes' if facing   else 'No',
            'Yes' if engaged  else 'No',
            round(float(row.get('Nose_X', 0)), 4)
                if 'Nose_X' in row and pd.notna(row.get('Nose_X')) else '—',
            round(float(row.get('Nose_Y', 0)), 4)
                if 'Nose_Y' in row and pd.notna(row.get('Nose_Y')) else '—',
        ]

        bool_cols = {3: dist_ok, 4: facing, 5: engaged}

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9)
            c.alignment = _align(h='center')
            c.border    = BORDER_THIN

            if col in bool_cols:
                ok = bool_cols[col]
                c.fill = _fill(CLR_GREEN_LIGHT if ok else CLR_RED_LIGHT)
                c.font = _font(bold=True, size=9,
                               color=CLR_GREEN_DARK if ok else CLR_RED_DARK)
            else:
                c.fill = _fill(bg)


# ---------------------------------------------------------------------------
# Sheet 5 — Disengagement Events
# ---------------------------------------------------------------------------
def _write_disengagement_events(wb, events):
    ws = wb.create_sheet('Disengagement Events')
    cols = 6
    _sheet_title(ws, 'Disengagement Events',
                 'Each row is one continuous period where the student was '
                 'fully disengaged (gaze off-screen AND too far AND not facing '
                 'forward — all three simultaneously).',
                 col_span=cols)

    headers = ['Event #', 'Start Time (s)', 'End Time (s)',
               'Duration (s)', 'Severity', 'Interpretation']
    _header_row(ws, 3, headers)
    _set_col_widths(ws, [10, 16, 14, 14, 14, 40])

    def _severity(dur):
        if dur < 2:
            return ('Brief', CLR_AMBER_LIGHT, CLR_AMBER_DARK)
        elif dur < 5:
            return ('Moderate', CLR_AMBER_LIGHT, CLR_AMBER_DARK)
        else:
            return ('Prolonged', CLR_RED_LIGHT, CLR_RED_DARK)

    def _interpretation(dur):
        if dur < 2:
            return 'Natural task-switching — not a concern'
        elif dur < 5:
            return 'Monitor — student may be distracted'
        else:
            return 'Significant disengagement — teacher attention recommended'

    if not events:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=cols)
        c = ws.cell(row=4, column=1,
                    value='No disengagement events recorded — excellent session!')
        c.font      = _font(bold=True, color=CLR_GREEN_DARK)
        c.fill      = _fill(CLR_GREEN_LIGHT)
        c.alignment = _align(h='center')
        return

    for i, event in enumerate(events):
        r   = i + 4
        dur = event.get('duration_s', 0)
        sev_label, sev_bg, sev_fg = _severity(dur)
        alt = (i % 2 == 1)
        bg  = CLR_GREY_ROW if alt else CLR_WHITE

        values = [
            i + 1,
            round(event.get('start_s', 0), 2),
            round(event.get('end_s', 0), 2),
            round(dur, 2),
            sev_label,
            _interpretation(dur),
        ]

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=10)
            c.alignment = _align(h='center' if col < 6 else 'left',
                                 wrap=(col == 6))
            c.border    = BORDER_THIN

            if col == 5:
                c.fill = _fill(sev_bg)
                c.font = _font(bold=True, size=10, color=sev_fg)
            else:
                c.fill = _fill(bg)

        ws.row_dimensions[r].height = 20

    # Totals row
    n = len(events) + 4
    ws.cell(row=n, column=1, value='Total events').font = _font(bold=True)
    ws.cell(row=n, column=1).fill = _fill(CLR_HEADER_LIGHT)
    ws.cell(row=n, column=2, value=len(events)).font = _font(bold=True)
    ws.cell(row=n, column=2).fill = _fill(CLR_HEADER_LIGHT)
    ws.cell(row=n, column=2).alignment = _align(h='center')

    total_dur = sum(e.get('duration_s', 0) for e in events)
    ws.cell(row=n, column=4, value=round(total_dur, 2)).font = _font(bold=True)
    ws.cell(row=n, column=4).fill = _fill(CLR_HEADER_LIGHT)
    ws.cell(row=n, column=4).alignment = _align(h='center')


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------
def export_session_xlsx(session_folder: str,
                        output_path: str = None) -> str:
    """
    Generate a teacher-friendly Excel workbook for a session.

    Parameters
    ----------
    session_folder : str
        Path to the session recording folder (contains tobii_gaze.csv,
        upper_body.csv, and analysis/ subfolder).
    output_path : str, optional
        Where to save the .xlsx file. Defaults to
        session_folder/analysis/session_report.xlsx

    Returns
    -------
    str — absolute path to the saved .xlsx file
    """
    session_folder = os.path.abspath(session_folder)
    analysis_dir   = os.path.join(session_folder, 'analysis')
    session_label  = os.path.basename(session_folder)

    # --- Load data ---
    summary_path = os.path.join(analysis_dir, 'session_summary.json')
    merged_path  = os.path.join(analysis_dir, 'engagement_merged.csv')
    score_path   = os.path.join(analysis_dir, 'engagement_score.csv')

    if not os.path.exists(summary_path):
        raise FileNotFoundError(
            f"No analysis found at: {summary_path}\n"
            "Run engagement_analysis.py on this session first.")

    with open(summary_path) as f:
        summary = json.load(f)

    merged_df = pd.read_csv(merged_path) if os.path.exists(merged_path) else pd.DataFrame()
    score_df  = pd.read_csv(score_path)  if os.path.exists(score_path)  else pd.DataFrame()

    # Fix boolean columns
    for col in ['on_screen', 'Distance_OK', 'Facing_Forward',
                'Body_Engaged', 'Engaged']:
        if col in merged_df.columns:
            merged_df[col] = (merged_df[col].astype(str).str.lower()
                              .map({'true': True, 'false': False,
                                    '1': True, '0': False})
                              .fillna(False))

    # --- Build workbook ---
    wb = Workbook()

    # Load session metadata if present
    meta_path = os.path.join(session_folder, 'session_metadata.json')
    session_metadata = None
    if os.path.exists(meta_path):
        try:
            with open(meta_path) as f:
                session_metadata = json.load(f)
            print(f"[META] Loaded metadata for report: {meta_path}")
        except Exception:
            pass

    _write_session_report(wb, summary, session_label, metadata=session_metadata)
    _write_engagement_summary(wb, score_df)
    _write_gaze_data(wb, merged_df)
    _write_body_data(wb, merged_df)
    _write_disengagement_events(wb, summary.get('events', []))

    # --- Freeze top rows on all sheets ---
    for ws in wb.worksheets:
        ws.freeze_panes = 'A4'
        ws.sheet_view.showGridLines = False

    # --- Save ---
    if output_path is None:
        os.makedirs(analysis_dir, exist_ok=True)
        output_path = os.path.join(analysis_dir, 'session_report.xlsx')

    output_path = os.path.abspath(output_path)
    wb.save(output_path)
    print(f"[EXPORT] Session report saved: {output_path}")
    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python session_exporter.py <session_folder> [output.xlsx]")
        sys.exit(1)
    folder = sys.argv[1]
    out    = sys.argv[2] if len(sys.argv) > 2 else None
    path   = export_session_xlsx(folder, out)
    print(f"Done: {path}")